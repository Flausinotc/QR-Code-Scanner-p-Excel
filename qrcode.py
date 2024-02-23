import cv2
from pyzbar.pyzbar import decode
import openpyxl
from openpyxl import Workbook

def read_qr_code():
    cap = cv2.VideoCapture(0)

    while True:
        ret, frame = cap.read()

        decoded_objs = decode(frame)

        for obj in decoded_objs:
            if obj.type == 'QRCODE':
                print("Valor do QR code:", obj.data.decode())
                cap.release()
                cv2.destroyAllWindows()
                return obj.data.decode()

        cv2.imshow("QR Code Scanner", frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

def export_to_excel(data):
    try:
        wb = openpyxl.load_workbook("qr_code_data.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active

        ws.append(["QR Code Data"])  

    next_row = ws.max_row + 1

    ws.cell(row=next_row, column=1).value = data

    wb.save("qr_code_data.xlsx")

if __name__ == "__main__":
    qr_code_data = read_qr_code()

    export_to_excel(qr_code_data)
